import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import numpy as np

def main():
    print("=== Access Requests Verification Report ===\n")
    
    # Define the file path
    file_path = r"D:\work\college_work\Coop_1\ops_work\cx034\CX034 - IPC - Part X Access Requests Completed - Diane_SJ200925.xlsx"
    
    # Check if file exists
    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        return
    
    print(f"📁 Processing file: {os.path.basename(file_path)}")
    
    try:
        # Load the workbook
        wb = load_workbook(file_path, data_only=True)
        print("✅ Workbook loaded successfully")
        
        # Check if required sheets exist
        sheet_names = wb.sheetnames
        print(f"📊 Sheets found: {', '.join(sheet_names)}")
        
        if 'Summary Page' not in sheet_names or 'Standard Report' not in sheet_names:
            print("❌ Required sheets not found!")
            return
        
        # Read Standard Report data using pandas for easier filtering
        std_report_df = pd.read_excel(file_path, sheet_name='Standard Report', header=1)
        print("✅ Standard Report data loaded")
        
        # Get column names for debugging
        print(f"\n📋 Standard Report columns: {list(std_report_df.columns)}")
        
        # Read Summary Page data
        summary_sheet = wb['Summary Page']
        
        # Extract values from rows 22-37, columns B, D, F, H
        summary_data = []
        for row in range(22, 38):
            summary_data.append({
                'row': row,
                'B': summary_sheet[f'B{row}'].value,
                'D': summary_sheet[f'D{row}'].value,
                'F': summary_sheet[f'F{row}'].value,
                'H': summary_sheet[f'H{row}'].value
            })
        
        # Display summary data
        print("\n📈 Summary Page Data (Rows 22-37):")
        print("-" * 80)
        print(f"{'Row':<4} | {'B (Clients)':<12} | {'D (Cases)':<10} | {'F (Participants)':<16} | {'H (Intake Cases)':<15}")
        print("-" * 80)
        for data in summary_data:
            print(f"{data['row']:<4} | {str(data['B']):<12} | {str(data['D']):<10} | {str(data['F']):<16} | {str(data['H']):<15}")
        
        # Perform verification for each row
        print("\n" + "="*80)
        print("VERIFICATION RESULTS")
        print("="*80)
        
        verification_results = []
        
        # Row 22: Full Access
        result_22 = verify_row_22(std_report_df, summary_data[0])
        verification_results.append(result_22)
        
        # Row 23: Partial Access - Part X Used To Deny Access
        result_23 = verify_row_23(std_report_df, summary_data[1])
        verification_results.append(result_23)
        
        # Row 24: Partial Access - Records Do Not Exist/Cannot Be Found
        result_24 = verify_row_24(std_report_df, summary_data[2])
        verification_results.append(result_24)
        
        # Row 25: Partial Access - Part X Does Not Apply
        result_25 = verify_row_25(std_report_df, summary_data[3])
        verification_results.append(result_25)
        
        # Row 26: Partial Access - Other
        result_26 = verify_row_26(std_report_df, summary_data[4])
        verification_results.append(result_26)
        
        # Row 27: Partial Access - Requestor's Information cannot reasonably be severed
        result_27 = verify_row_27(std_report_df, summary_data[5])
        verification_results.append(result_27)
        
        # Row 28: No Information Released - Part X Used To Deny Access
        result_28 = verify_row_28(std_report_df, summary_data[6])
        verification_results.append(result_28)
        
        # Row 29: No Information Released - Records Do Not Exist/Cannot Be Found
        result_29 = verify_row_29(std_report_df, summary_data[7])
        verification_results.append(result_29)
        
        # Row 30: No Information Released - Part X Does Not Apply
        result_30 = verify_row_30(std_report_df, summary_data[8])
        verification_results.append(result_30)
        
        # Row 31: No Information Released - Other
        result_31 = verify_row_31(std_report_df, summary_data[9])
        verification_results.append(result_31)
        
        # Row 32: No Information Released - Requestor's Information cannot reasonably be severed
        result_32 = verify_row_32(std_report_df, summary_data[10])
        verification_results.append(result_32)
        
        # Row 33: No Information Released - Intake Case Only
        result_33 = verify_row_33(std_report_df, summary_data[11])
        verification_results.append(result_33)
        
        # Row 34: Withdrawn or Abandoned
        result_34 = verify_row_34(std_report_df, summary_data[12])
        verification_results.append(result_34)
        
        # Row 35: Documentation Completed
        result_35 = verify_row_35(std_report_df, summary_data[13])
        verification_results.append(result_35)
        
        # Row 36: Total Distinct Outcomes of Access Requests
        result_36 = verify_row_36(verification_results, summary_data[14])
        verification_results.append(result_36)
        
        # Row 37: Total Distinct Outcomes - Partial or No Information Released - Part X Used To Deny Access
        result_37 = verify_row_37(verification_results, summary_data[15])
        verification_results.append(result_37)
        
        # Generate final report
        generate_final_report(verification_results)
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error processing file: {str(e)}")
        import traceback
        traceback.print_exc()

def verify_row_22(df, summary_row):
    """Verify Full Access"""
    print(f"\n🔍 Verifying Row {summary_row['row']}: Full Access")
    
    # Map column names based on your data structure
    closure_reason_col = 'Personal Information Maximum Submitted Closure Reason'
    case_num_col = 'Personal Information Case #'
    clients_col = 'Personal Information Number of Clients'
    
    if not all(col in df.columns for col in [closure_reason_col, case_num_col, clients_col]):
        return {'row': summary_row['row'], 'status': 'ERROR', 'message': 'Required columns not found'}
    
    # Filter for Full Access
    filtered_df = df[df[closure_reason_col] == 'Full Access']
    
    # Calculate expected values
    expected_clients = filtered_df[clients_col].sum()
    expected_cases = filtered_df[case_num_col].nunique()
    
    actual_clients = summary_row['B']
    actual_cases = summary_row['D']
    
    clients_match = expected_clients == actual_clients
    cases_match = expected_cases == actual_cases
    
    return {
        'row': summary_row['row'],
        'category': 'Full Access',
        'expected_B': expected_clients,
        'actual_B': actual_clients,
        'B_match': clients_match,
        'expected_D': expected_cases,
        'actual_D': actual_cases,
        'D_match': cases_match,
        'status': 'PASS' if clients_match and cases_match else 'FAIL'
    }

def verify_row_23(df, summary_row):
    """Verify Partial Access - Part X Used To Deny Access"""
    print(f"\n🔍 Verifying Row {summary_row['row']}: Partial Access - Part X Deny")
    
    closure_reason_col = 'Personal Information Maximum Submitted Closure Reason'
    provisions_col = 'Personal Information Provisions to Deny'
    case_num_col = 'Personal Information Case #'
    clients_col = 'Personal Information Number of Clients'
    
    part_x_denial_provisions = [
        'Refusal – s.312(1)(a) Legal privilege',
        'Refusal – s.312(1)(b) Prohibited by law or court order',
        'Refusal – s.312(1)(c) Collected for legal proceeding which has not concluded',
        'Refusal – s.312(1)(d)(i) Risk of serious harm',
        'Refusal – s.312(1)(d)(ii) Identifies individual with duty to report',
        'Refusal – s.312(1)(d)(iii) Identifies information sources who reported in confidence',
        'Refusal – s.314(6) Frivolous or vexatious or bad faith (access)'
    ]
    
    # Filter for Partial Access and check if ANY Part X provision appears in the string
    filtered_df = df[
        (df[closure_reason_col] == 'Partial Access')
    ]
    
    # Check if any Part X provision is contained in the provisions string
    def contains_part_x_denial(provisions_str):
        if pd.isna(provisions_str):
            return False
        provisions_str = str(provisions_str)
        return any(provision in provisions_str for provision in part_x_denial_provisions)
    
    filtered_df = filtered_df[filtered_df[provisions_col].apply(contains_part_x_denial)]
    
    expected_clients = filtered_df[clients_col].sum()
    expected_cases = filtered_df[case_num_col].nunique()
    
    actual_clients = summary_row['B']
    actual_cases = summary_row['D']
    
    clients_match = expected_clients == actual_clients
    cases_match = expected_cases == actual_cases
    
    # Debug output
    print(f"   Found {len(filtered_df)} cases with Part X denials")
    if len(filtered_df) > 0:
        print(f"   Sample provisions: {filtered_df[provisions_col].head(3).tolist()}")
    
    return {
        'row': summary_row['row'],
        'category': 'Partial Access - Part X Deny',
        'expected_B': expected_clients,
        'actual_B': actual_clients,
        'B_match': clients_match,
        'expected_D': expected_cases,
        'actual_D': actual_cases,
        'D_match': cases_match,
        'status': 'PASS' if clients_match and cases_match else 'FAIL'
    }

def verify_row_24(df, summary_row):
    """Verify Partial Access - Records Do Not Exist/Cannot Be Found"""
    print(f"\n🔍 Verifying Row {summary_row['row']}: Partial Access - Records Not Found")
    
    closure_reason_col = 'Personal Information Maximum Submitted Closure Reason'
    provisions_col = 'Personal Information Provisions to Deny'
    case_num_col = 'Personal Information Case #'
    clients_col = 'Personal Information Number of Clients'
    
    # Filter for Partial Access and check if records not found appears
    filtered_df = df[
        (df[closure_reason_col] == 'Partial Access')
    ]
    
    def contains_records_not_found(provisions_str):
        if pd.isna(provisions_str):
            return False
        provisions_str = str(provisions_str)
        return 'records do not exist or cannot be found' in provisions_str or 'No record exists, or none found' in provisions_str
    
    filtered_df = filtered_df[filtered_df[provisions_col].apply(contains_records_not_found)]
    
    expected_clients = filtered_df[clients_col].sum()
    expected_cases = filtered_df[case_num_col].nunique()
    
    actual_clients = summary_row['B']
    actual_cases = summary_row['D']
    
    clients_match = expected_clients == actual_clients
    cases_match = expected_cases == actual_cases
    
    return {
        'row': summary_row['row'],
        'category': 'Partial Access - Records Not Found',
        'expected_B': expected_clients,
        'actual_B': actual_clients,
        'B_match': clients_match,
        'expected_D': expected_cases,
        'actual_D': actual_cases,
        'D_match': cases_match,
        'status': 'PASS' if clients_match and cases_match else 'FAIL'
    }

def verify_row_25(df, summary_row):
    """Verify Partial Access - Part X Does Not Apply"""
    print(f"\n🔍 Verifying Row {summary_row['row']}: Partial Access - Part X Does Not Apply")
    
    closure_reason_col = 'Personal Information Maximum Submitted Closure Reason'
    provisions_col = 'Personal Information Provisions to Deny'
    case_num_col = 'Personal Information Case #'
    clients_col = 'Personal Information Number of Clients'
    
    # Filter for Partial Access and check if "Part X does not apply" appears
    filtered_df = df[
        (df[closure_reason_col] == 'Partial Access')
    ]
    
    def contains_part_x_does_not_apply(provisions_str):
        if pd.isna(provisions_str):
            return False
        provisions_str = str(provisions_str)
        return 'Part X does not apply' in provisions_str
    
    filtered_df = filtered_df[filtered_df[provisions_col].apply(contains_part_x_does_not_apply)]
    
    expected_clients = filtered_df[clients_col].sum()
    expected_cases = filtered_df[case_num_col].nunique()
    
    actual_clients = summary_row['B']
    actual_cases = summary_row['D']
    
    clients_match = expected_clients == actual_clients
    cases_match = expected_cases == actual_cases
    
    return {
        'row': summary_row['row'],
        'category': 'Partial Access - Part X Does Not Apply',
        'expected_B': expected_clients,
        'actual_B': actual_clients,
        'B_match': clients_match,
        'expected_D': expected_cases,
        'actual_D': actual_cases,
        'D_match': cases_match,
        'status': 'PASS' if clients_match and cases_match else 'FAIL'
    }

def verify_row_26(df, summary_row):
    """Verify Partial Access - Other"""
    print(f"\n🔍 Verifying Row {summary_row['row']}: Partial Access - Other")
    
    closure_reason_col = 'Personal Information Maximum Submitted Closure Reason'
    provisions_col = 'Personal Information Provisions to Deny'
    case_num_col = 'Personal Information Case #'
    clients_col = 'Personal Information Number of Clients'
    
    # Filter for Partial Access and "Other" but NOT containing other specific provisions
    filtered_df = df[
        (df[closure_reason_col] == 'Partial Access')
    ]
    
    def is_other_only(provisions_str):
        if pd.isna(provisions_str):
            return False
        
        provisions_str = str(provisions_str)
        # Should contain "Other" but NOT contain Part X provisions or cannot be severed
        has_other = 'Refusal - Other' in provisions_str
        
        # Should NOT contain other specific refusal reasons
        has_other_refusals = any(refusal in provisions_str for refusal in [
            's.312(1)(a)', 's.312(1)(b)', 's.312(1)(c)', 's.312(1)(d)',
            's.314(6)', 'Part X does not apply', 'cannot reasonably be severed',
            'records do not exist', 'No record exists'
        ])
        
        return has_other and not has_other_refusals
    
    filtered_df = filtered_df[filtered_df[provisions_col].apply(is_other_only)]
    
    expected_clients = filtered_df[clients_col].sum()
    expected_cases = filtered_df[case_num_col].nunique()
    
    actual_clients = summary_row['B']
    actual_cases = summary_row['D']
    
    clients_match = expected_clients == actual_clients
    cases_match = expected_cases == actual_cases
    
    return {
        'row': summary_row['row'],
        'category': 'Partial Access - Other',
        'expected_B': expected_clients,
        'actual_B': actual_clients,
        'B_match': clients_match,
        'expected_D': expected_cases,
        'actual_D': actual_cases,
        'D_match': cases_match,
        'status': 'PASS' if clients_match and cases_match else 'FAIL'
    }

def verify_row_27(df, summary_row):
    """Verify Partial Access - Requestor's Information cannot reasonably be severed"""
    print(f"\n🔍 Verifying Row {summary_row['row']}: Partial Access - Cannot Be Severed")
    
    closure_reason_col = 'Personal Information Maximum Submitted Closure Reason'
    provisions_col = 'Personal Information Provisions to Deny'
    case_num_col = 'Personal Information Case #'
    clients_col = 'Personal Information Number of Clients'
    
    # Filter for Partial Access and check if "cannot reasonably be severed" appears
    filtered_df = df[
        (df[closure_reason_col] == 'Partial Access')
    ]
    
    def contains_cannot_be_severed(provisions_str):
        if pd.isna(provisions_str):
            return False
        provisions_str = str(provisions_str)
        return 'cannot reasonably be severed' in provisions_str
    
    filtered_df = filtered_df[filtered_df[provisions_col].apply(contains_cannot_be_severed)]
    
    expected_clients = filtered_df[clients_col].sum()
    expected_cases = filtered_df[case_num_col].nunique()
    
    actual_clients = summary_row['B']
    actual_cases = summary_row['D']
    
    clients_match = expected_clients == actual_clients
    cases_match = expected_cases == actual_cases
    
    return {
        'row': summary_row['row'],
        'category': 'Partial Access - Cannot Be Severed',
        'expected_B': expected_clients,
        'actual_B': actual_clients,
        'B_match': clients_match,
        'expected_D': expected_cases,
        'actual_D': actual_cases,
        'D_match': cases_match,
        'status': 'PASS' if clients_match and cases_match else 'FAIL'
    }

def verify_row_28(df, summary_row):
    """Verify No Information Released - Part X Used To Deny Access"""
    print(f"\n🔍 Verifying Row {summary_row['row']}: No Info Released - Part X Deny")
    
    closure_reason_col = 'Personal Information Maximum Submitted Closure Reason'
    provisions_col = 'Personal Information Provisions to Deny'
    case_num_col = 'Personal Information Case #'
    clients_col = 'Personal Information Number of Clients'
    
    part_x_denial_provisions = [
        'Refusal – s.312(1)(a) Legal privilege',
        'Refusal – s.312(1)(b) Prohibited by law or court order',
        'Refusal – s.312(1)(c) Collected for legal proceeding which has not concluded',
        'Refusal – s.312(1)(d)(i) Risk of serious harm',
        'Refusal – s.312(1)(d)(ii) Identifies individual with duty to report',
        'Refusal – s.312(1)(d)(iii) Identifies information sources who reported in confidence',
        'Refusal – s.314(6) Frivolous or vexatious or bad faith (access)'
    ]
    
    # Filter for No Information Released and check if ANY Part X provision appears
    filtered_df = df[
        (df[closure_reason_col] == 'No Information Released')
    ]
    
    def contains_part_x_denial(provisions_str):
        if pd.isna(provisions_str):
            return False
        provisions_str = str(provisions_str)
        return any(provision in provisions_str for provision in part_x_denial_provisions)
    
    filtered_df = filtered_df[filtered_df[provisions_col].apply(contains_part_x_denial)]
    
    expected_clients = filtered_df[clients_col].sum()
    expected_cases = filtered_df[case_num_col].nunique()
    
    actual_clients = summary_row['B']
    actual_cases = summary_row['D']
    
    clients_match = expected_clients == actual_clients
    cases_match = expected_cases == actual_cases
    
    return {
        'row': summary_row['row'],
        'category': 'No Info Released - Part X Deny',
        'expected_B': expected_clients,
        'actual_B': actual_clients,
        'B_match': clients_match,
        'expected_D': expected_cases,
        'actual_D': actual_cases,
        'D_match': cases_match,
        'status': 'PASS' if clients_match and cases_match else 'FAIL'
    }

def verify_row_29(df, summary_row):
    """Verify No Information Released - Records Do Not Exist/Cannot Be Found"""
    print(f"\n🔍 Verifying Row {summary_row['row']}: No Info Released - Records Not Found")
    
    closure_reason_col = 'Personal Information Maximum Submitted Closure Reason'
    provisions_col = 'Personal Information Provisions to Deny'
    case_num_col = 'Personal Information Case #'
    clients_col = 'Personal Information Number of Clients'
    
    # Filter for No Information Released and check if records not found appears
    filtered_df = df[
        (df[closure_reason_col] == 'No Information Released')
    ]
    
    def contains_records_not_found(provisions_str):
        if pd.isna(provisions_str):
            return False
        provisions_str = str(provisions_str)
        return 'No record exists, or none found' in provisions_str
    
    filtered_df = filtered_df[filtered_df[provisions_col].apply(contains_records_not_found)]
    
    expected_clients = filtered_df[clients_col].sum()
    expected_cases = filtered_df[case_num_col].nunique()
    
    actual_clients = summary_row['B']
    actual_cases = summary_row['D']
    
    clients_match = expected_clients == actual_clients
    cases_match = expected_cases == actual_cases
    
    return {
        'row': summary_row['row'],
        'category': 'No Info Released - Records Not Found',
        'expected_B': expected_clients,
        'actual_B': actual_clients,
        'B_match': clients_match,
        'expected_D': expected_cases,
        'actual_D': actual_cases,
        'D_match': cases_match,
        'status': 'PASS' if clients_match and cases_match else 'FAIL'
    }

def verify_row_30(df, summary_row):
    """Verify No Information Released - Part X Does Not Apply"""
    print(f"\n🔍 Verifying Row {summary_row['row']}: No Info Released - Part X Does Not Apply")
    
    closure_reason_col = 'Personal Information Maximum Submitted Closure Reason'
    provisions_col = 'Personal Information Provisions to Deny'
    case_num_col = 'Personal Information Case #'
    clients_col = 'Personal Information Number of Clients'
    
    # Filter for No Information Released and check if "Part X does not apply" appears
    filtered_df = df[
        (df[closure_reason_col] == 'No Information Released')
    ]
    
    def contains_part_x_does_not_apply(provisions_str):
        if pd.isna(provisions_str):
            return False
        provisions_str = str(provisions_str)
        return 'Part X does not apply' in provisions_str
    
    filtered_df = filtered_df[filtered_df[provisions_col].apply(contains_part_x_does_not_apply)]
    
    expected_clients = filtered_df[clients_col].sum()
    expected_cases = filtered_df[case_num_col].nunique()
    
    actual_clients = summary_row['B']
    actual_cases = summary_row['D']
    
    clients_match = expected_clients == actual_clients
    cases_match = expected_cases == actual_cases
    
    return {
        'row': summary_row['row'],
        'category': 'No Info Released - Part X Does Not Apply',
        'expected_B': expected_clients,
        'actual_B': actual_clients,
        'B_match': clients_match,
        'expected_D': expected_cases,
        'actual_D': actual_cases,
        'D_match': cases_match,
        'status': 'PASS' if clients_match and cases_match else 'FAIL'
    }

def verify_row_31(df, summary_row):
    """Verify No Information Released - Other"""
    print(f"\n🔍 Verifying Row {summary_row['row']}: No Info Released - Other")
    
    closure_reason_col = 'Personal Information Maximum Submitted Closure Reason'
    provisions_col = 'Personal Information Provisions to Deny'
    case_num_col = 'Personal Information Case #'
    clients_col = 'Personal Information Number of Clients'
    
    # Filter for No Information Released and "Other" but NOT containing other specific provisions
    filtered_df = df[
        (df[closure_reason_col] == 'No Information Released')
    ]
    
    def is_other_only(provisions_str):
        if pd.isna(provisions_str):
            return False
        
        provisions_str = str(provisions_str)
        # Should contain "Other" but NOT contain Part X provisions or cannot be severed
        has_other = 'Refusal - Other' in provisions_str
        
        # Should NOT contain other specific refusal reasons
        has_other_refusals = any(refusal in provisions_str for refusal in [
            's.312(1)(a)', 's.312(1)(b)', 's.312(1)(c)', 's.312(1)(d)',
            's.314(6)', 'Part X does not apply', 'cannot reasonably be severed',
            'No record exists'
        ])
        
        return has_other and not has_other_refusals
    
    filtered_df = filtered_df[filtered_df[provisions_col].apply(is_other_only)]
    
    expected_clients = filtered_df[clients_col].sum()
    expected_cases = filtered_df[case_num_col].nunique()
    
    actual_clients = summary_row['B']
    actual_cases = summary_row['D']
    
    clients_match = expected_clients == actual_clients
    cases_match = expected_cases == actual_cases
    
    return {
        'row': summary_row['row'],
        'category': 'No Info Released - Other',
        'expected_B': expected_clients,
        'actual_B': actual_clients,
        'B_match': clients_match,
        'expected_D': expected_cases,
        'actual_D': actual_cases,
        'D_match': cases_match,
        'status': 'PASS' if clients_match and cases_match else 'FAIL'
    }

def verify_row_32(df, summary_row):
    """Verify No Information Released - Requestor's Information cannot reasonably be severed"""
    print(f"\n🔍 Verifying Row {summary_row['row']}: No Info Released - Cannot Be Severed")
    
    closure_reason_col = 'Personal Information Maximum Submitted Closure Reason'
    provisions_col = 'Personal Information Provisions to Deny'
    case_num_col = 'Personal Information Case #'
    clients_col = 'Personal Information Number of Clients'
    
    # Filter for No Information Released and check if "cannot reasonably be severed" appears
    filtered_df = df[
        (df[closure_reason_col] == 'No Information Released')
    ]
    
    def contains_cannot_be_severed(provisions_str):
        if pd.isna(provisions_str):
            return False
        provisions_str = str(provisions_str)
        return 'cannot reasonably be severed' in provisions_str
    
    filtered_df = filtered_df[filtered_df[provisions_col].apply(contains_cannot_be_severed)]
    
    expected_clients = filtered_df[clients_col].sum()
    expected_cases = filtered_df[case_num_col].nunique()
    
    actual_clients = summary_row['B']
    actual_cases = summary_row['D']
    
    clients_match = expected_clients == actual_clients
    cases_match = expected_cases == actual_cases
    
    return {
        'row': summary_row['row'],
        'category': 'No Info Released - Cannot Be Severed',
        'expected_B': expected_clients,
        'actual_B': actual_clients,
        'B_match': clients_match,
        'expected_D': expected_cases,
        'actual_D': actual_cases,
        'D_match': cases_match,
        'status': 'PASS' if clients_match and cases_match else 'FAIL'
    }

def verify_row_33(df, summary_row):
    """Verify No Information Released - Intake Case Only"""
    print(f"\n🔍 Verifying Row {summary_row['row']}: No Information Released - Intake Only")
    
    intake_disposition_col = 'Intake Maximum Submitted  Disposition'
    intake_participants_col = 'Intake Number of Participants'
    intake_case_col = 'Intake Case #'
    
    filtered_df = df[df[intake_disposition_col] == 'No Information Released']
    
    expected_participants = filtered_df[intake_participants_col].sum()
    expected_cases = filtered_df[intake_case_col].nunique()
    
    actual_participants = summary_row['F']
    actual_cases = summary_row['H']
    
    participants_match = expected_participants == actual_participants
    cases_match = expected_cases == actual_cases
    
    return {
        'row': summary_row['row'],
        'category': 'No Information Released - Intake Only',
        'expected_F': expected_participants,
        'actual_F': actual_participants,
        'F_match': participants_match,
        'expected_H': expected_cases,
        'actual_H': actual_cases,
        'H_match': cases_match,
        'status': 'PASS' if participants_match and cases_match else 'FAIL'
    }

def verify_row_34(df, summary_row):
    """Verify Withdrawn or Abandoned"""
    print(f"\n🔍 Verifying Row {summary_row['row']}: Withdrawn or Abandoned")
    
    closure_reason_col = 'Personal Information Maximum Submitted Closure Reason'
    case_num_col = 'Personal Information Case #'
    clients_col = 'Personal Information Number of Clients'
    intake_disposition_col = 'Intake Maximum Submitted  Disposition'
    intake_participants_col = 'Intake Number of Participants'
    intake_case_col = 'Intake Case #'
    
    # Personal Information Cases
    pi_filtered_df = df[df[closure_reason_col] == 'Request withdrawn or abandoned']
    expected_pi_clients = pi_filtered_df[clients_col].sum()
    expected_pi_cases = pi_filtered_df[case_num_col].nunique()
    
    # Intake Cases
    intake_filtered_df = df[df[intake_disposition_col] == 'Request withdrawn or abandoned']
    expected_intake_participants = intake_filtered_df[intake_participants_col].sum()
    expected_intake_cases = intake_filtered_df[intake_case_col].nunique()
    
    actual_pi_clients = summary_row['B']
    actual_pi_cases = summary_row['D']
    actual_intake_participants = summary_row['F']
    actual_intake_cases = summary_row['H']
    
    pi_clients_match = expected_pi_clients == actual_pi_clients
    pi_cases_match = expected_pi_cases == actual_pi_cases
    intake_participants_match = expected_intake_participants == actual_intake_participants
    intake_cases_match = expected_intake_cases == actual_intake_cases
    
    return {
        'row': summary_row['row'],
        'category': 'Withdrawn or Abandoned',
        'expected_B': expected_pi_clients,
        'actual_B': actual_pi_clients,
        'B_match': pi_clients_match,
        'expected_D': expected_pi_cases,
        'actual_D': actual_pi_cases,
        'D_match': pi_cases_match,
        'expected_F': expected_intake_participants,
        'actual_F': actual_intake_participants,
        'F_match': intake_participants_match,
        'expected_H': expected_intake_cases,
        'actual_H': actual_intake_cases,
        'H_match': intake_cases_match,
        'status': 'PASS' if all([pi_clients_match, pi_cases_match, intake_participants_match, intake_cases_match]) else 'FAIL'
    }

def verify_row_35(df, summary_row):
    """Verify Documentation Completed"""
    print(f"\n🔍 Verifying Row {summary_row['row']}: Documentation Completed")
    
    closure_reason_col = 'Personal Information Maximum Submitted Closure Reason'
    case_num_col = 'Personal Information Case #'
    clients_col = 'Personal Information Number of Clients'
    
    filtered_df = df[df[closure_reason_col] == 'Documentation Completed']
    
    expected_clients = filtered_df[clients_col].sum()
    expected_cases = filtered_df[case_num_col].nunique()
    
    actual_clients = summary_row['B']
    actual_cases = summary_row['D']
    
    clients_match = expected_clients == actual_clients
    cases_match = expected_cases == actual_cases
    
    return {
        'row': summary_row['row'],
        'category': 'Documentation Completed',
        'expected_B': expected_clients,
        'actual_B': actual_clients,
        'B_match': clients_match,
        'expected_D': expected_cases,
        'actual_D': actual_cases,
        'D_match': cases_match,
        'status': 'PASS' if clients_match and cases_match else 'FAIL'
    }

def verify_row_36(results, summary_row):
    """Verify Total Distinct Outcomes (Sum of Rows 22-35)"""
    print(f"\n🔍 Verifying Row {summary_row['row']}: Total Distinct Outcomes")
    
    # Sum EXPECTED values from rows 22-35 (index 0-13 in results)
    relevant_results = results[0:14]  # Rows 22-35
    
    # Use EXPECTED values, not actual values
    total_B = sum(r.get('expected_B', 0) for r in relevant_results if r.get('expected_B') not in [None, 'N/A'])
    total_D = sum(r.get('expected_D', 0) for r in relevant_results if r.get('expected_D') not in [None, 'N/A'])
    total_F = sum(r.get('expected_F', 0) for r in relevant_results if r.get('expected_F') not in [None, 'N/A'])
    total_H = sum(r.get('expected_H', 0) for r in relevant_results if r.get('expected_H') not in [None, 'N/A'])
    
    B_match = total_B == summary_row['B']
    D_match = total_D == summary_row['D']
    F_match = total_F == summary_row['F']
    H_match = total_H == summary_row['H']
    
    # Debug output to see what's being summed
    print(f"   Summing expected values from rows 22-35:")
    for i, result in enumerate(relevant_results):
        row_num = 22 + i
        print(f"   Row {row_num}: B={result.get('expected_B', 0)}, D={result.get('expected_D', 0)}, F={result.get('expected_F', 'N/A')}, H={result.get('expected_H', 'N/A')}")
    print(f"   Total Expected: B={total_B}, D={total_D}, F={total_F}, H={total_H}")
    print(f"   Actual Summary: B={summary_row['B']}, D={summary_row['D']}, F={summary_row['F']}, H={summary_row['H']}")
    
    return {
        'row': summary_row['row'],
        'category': 'Total Distinct Outcomes',
        'expected_B': total_B,
        'actual_B': summary_row['B'],
        'B_match': B_match,
        'expected_D': total_D,
        'actual_D': summary_row['D'],
        'D_match': D_match,
        'expected_F': total_F,
        'actual_F': summary_row['F'],
        'F_match': F_match,
        'expected_H': total_H,
        'actual_H': summary_row['H'],
        'H_match': H_match,
        'status': 'PASS' if all([B_match, D_match, F_match, H_match]) else 'FAIL'
    }

def verify_row_37(results, summary_row):
    """Verify Total Partial or No Info Released - Part X Deny (Row 23 + Row 28)"""
    print(f"\n🔍 Verifying Row {summary_row['row']}: Partial/No Info - Part X Deny")
    
    # Get EXPECTED values from row 23 (index 1) and row 28 (index 6)
    row_23 = results[1]  # Partial Access - Part X Deny
    row_28 = results[6]  # No Information Released - Part X Deny
    
    # Use EXPECTED values, not actual values
    total_B = (row_23.get('expected_B', 0) if row_23.get('expected_B') not in [None, 'N/A'] else 0) + \
              (row_28.get('expected_B', 0) if row_28.get('expected_B') not in [None, 'N/A'] else 0)
    total_D = (row_23.get('expected_D', 0) if row_23.get('expected_D') not in [None, 'N/A'] else 0) + \
              (row_28.get('expected_D', 0) if row_28.get('expected_D') not in [None, 'N/A'] else 0)
    
    B_match = total_B == summary_row['B']
    D_match = total_D == summary_row['D']
    
    # Debug output
    print(f"   Row 23 Expected: B={row_23.get('expected_B', 0)}, D={row_23.get('expected_D', 0)}")
    print(f"   Row 28 Expected: B={row_28.get('expected_B', 0)}, D={row_28.get('expected_D', 0)}")
    print(f"   Total Expected: B={total_B}, D={total_D}")
    print(f"   Actual Summary: B={summary_row['B']}, D={summary_row['D']}")
    
    return {
        'row': summary_row['row'],
        'category': 'Partial/No Info - Part X Deny',
        'expected_B': total_B,
        'actual_B': summary_row['B'],
        'B_match': B_match,
        'expected_D': total_D,
        'actual_D': summary_row['D'],
        'D_match': D_match,
        'status': 'PASS' if B_match and D_match else 'FAIL'
    }

def generate_final_report(results):
    """Generate comprehensive verification report"""
    print("\n" + "="*80)
    print("FINAL VERIFICATION REPORT")
    print("="*80)
    
    passed = 0
    failed = 0
    errors = 0
    
    for result in results:
        if result['status'] == 'PASS':
            passed += 1
            status_icon = '✅'
        elif result['status'] == 'FAIL':
            failed += 1
            status_icon = '❌'
        else:
            errors += 1
            status_icon = '⚠️'
        
        print(f"\n{status_icon} Row {result['row']}: {result['category']}")
        print(f"   Status: {result['status']}")
        
        if 'B_match' in result:
            b_icon = '✅' if result['B_match'] else '❌'
            print(f"   {b_icon} Clients: Expected {result.get('expected_B', 'N/A')} | Actual {result.get('actual_B', 'N/A')}")
        
        if 'D_match' in result:
            d_icon = '✅' if result['D_match'] else '❌'
            print(f"   {d_icon} Cases: Expected {result.get('expected_D', 'N/A')} | Actual {result.get('actual_D', 'N/A')}")
        
        if 'F_match' in result:
            f_icon = '✅' if result['F_match'] else '❌'
            print(f"   {f_icon} Participants: Expected {result.get('expected_F', 'N/A')} | Actual {result.get('actual_F', 'N/A')}")
        
        if 'H_match' in result:
            h_icon = '✅' if result['H_match'] else '❌'
            print(f"   {h_icon} Intake Cases: Expected {result.get('expected_H', 'N/A')} | Actual {result.get('actual_H', 'N/A')}")
    
    print("\n" + "="*80)
    print(f"SUMMARY: {passed} PASSED, {failed} FAILED, {errors} ERRORS")
    print(f"SUCCESS RATE: {(passed/(passed+failed+errors))*100:.1f}%")
    print("="*80)

if __name__ == "__main__":
    main()

'''
=== Access Requests Verification Report ===

📁 Processing file: CX034 - IPC - Part X Access Requests Completed - Diane_SJ200925.xlsx
✅ Workbook loaded successfully
📊 Sheets found: Cover Page, Standard Report, Summary Page
✅ Standard Report data loaded

📋 Standard Report columns: ['Team', 'Team Supervisor Last Name', 'Team Supervisor First Name', 'Case Supervisor Last Name', 'Case Supervisor First Name', 'Case Owner Last Name', 'Case Owner First Name', 'Secondary Case Owner Last Name', 'Secondary Case Owner First Name', 'Primary Client Last Name', 'Primary Client First Name', 'Primary Client #', 'Intake Case #', 'Intake Received Date', 'Intake Number of Participants', 'Primary Eligibility Spectrum Code', 'Secondary Eligibility Spectrum Code', 'Eligibility Spectrum Code Version', 'Missing From Primary ESC', 'Intake Maximum Submitted  Disposition', 'Intake Maximum Submitted Closure Reason', 'Intake Maximum Submitted Date', 'Intake Approved Date', 'Personal Information Case #', 'Personal Information Start Date', 'Personal Information Number of Clients', 'Intake Submission Date + 30 Days', 'Milestone Name', 'Milestone Expected End Date', 'Milestone Actual End Date', 'Waiver Request Date', 'Waiver Request Reason', 'Waiver Requested Expected End Date', 'Waiver Request Status', 'Invalid Waiver Extension Date', 'Personal Information Maximum Submitted for Case Closure Date', 'Personal Information Maximum Submitted Closure Reason', 'Personal Information Provisions to Deny']

📈 Summary Page Data (Rows 22-37):
--------------------------------------------------------------------------------
Row  | B (Clients)  | D (Cases)  | F (Participants) | H (Intake Cases)
--------------------------------------------------------------------------------
22   | 51           | 50         | N/A              | N/A
23   | 114          | 75         | N/A              | N/A
24   | 0            | 0          | N/A              | N/A
25   | 13           | 8          | N/A              | N/A
26   | 12           | 8          | N/A              | N/A
27   | 114          | 75         | N/A              | N/A
28   | 0            | 0          | N/A              | N/A
29   | 5            | 5          | N/A              | N/A
30   | 9            | 6          | N/A              | N/A
31   | 5            | 4          | N/A              | N/A
32   | 0            | 0          | N/A              | N/A
33   | N/A          | N/A        | 89               | 62
34   | 52           | 20         | 85               | 46
35   | 131          | 92         | N/A              | N/A
36   | 506          | 343        | 174              | 108
37   | 114          | 75         | N/A              | N/A

================================================================================
VERIFICATION RESULTS
================================================================================

🔍 Verifying Row 22: Full Access

🔍 Verifying Row 23: Partial Access - Part X Deny
   Found 0 cases with Part X denials

🔍 Verifying Row 24: Partial Access - Records Not Found

🔍 Verifying Row 25: Partial Access - Part X Does Not Apply

🔍 Verifying Row 26: Partial Access - Other

🔍 Verifying Row 27: Partial Access - Cannot Be Severed

🔍 Verifying Row 28: No Info Released - Part X Deny

🔍 Verifying Row 29: No Info Released - Records Not Found

🔍 Verifying Row 30: No Info Released - Part X Does Not Apply

🔍 Verifying Row 31: No Info Released - Other

🔍 Verifying Row 32: No Info Released - Cannot Be Severed

🔍 Verifying Row 33: No Information Released - Intake Only

🔍 Verifying Row 34: Withdrawn or Abandoned

🔍 Verifying Row 35: Documentation Completed

🔍 Verifying Row 36: Total Distinct Outcomes
   Summing expected values from rows 22-35:
   Row 22: B=51.0, D=50, F=N/A, H=N/A
   Row 23: B=0.0, D=0, F=N/A, H=N/A
   Row 24: B=0.0, D=0, F=N/A, H=N/A
   Row 25: B=13.0, D=8, F=N/A, H=N/A
   Row 26: B=4.0, D=2, F=N/A, H=N/A
   Row 27: B=114.0, D=75, F=N/A, H=N/A
   Row 28: B=0.0, D=0, F=N/A, H=N/A
   Row 29: B=5.0, D=5, F=N/A, H=N/A
   Row 30: B=9.0, D=6, F=N/A, H=N/A
   Row 31: B=4.0, D=3, F=N/A, H=N/A
   Row 32: B=0.0, D=0, F=N/A, H=N/A
   Row 33: B=0, D=0, F=89.0, H=62
   Row 34: B=52.0, D=20, F=85.0, H=46
   Row 35: B=131.0, D=92, F=N/A, H=N/A
   Total Expected: B=383.0, D=261, F=174.0, H=108
   Actual Summary: B=506, D=343, F=174, H=108

🔍 Verifying Row 37: Partial/No Info - Part X Deny
   Row 23 Expected: B=0.0, D=0
   Row 28 Expected: B=0.0, D=0
   Total Expected: B=0.0, D=0
   Actual Summary: B=114, D=75

================================================================================
FINAL VERIFICATION REPORT
================================================================================

✅ Row 22: Full Access
   Status: PASS
   ✅ Clients: Expected 51.0 | Actual 51
   ✅ Cases: Expected 50 | Actual 50

❌ Row 23: Partial Access - Part X Deny
   Status: FAIL
   ❌ Clients: Expected 0.0 | Actual 114
   ❌ Cases: Expected 0 | Actual 75

✅ Row 24: Partial Access - Records Not Found
   Status: PASS
   ✅ Clients: Expected 0.0 | Actual 0
   ✅ Cases: Expected 0 | Actual 0

✅ Row 25: Partial Access - Part X Does Not Apply
   Status: PASS
   ✅ Clients: Expected 13.0 | Actual 13
   ✅ Cases: Expected 8 | Actual 8

❌ Row 26: Partial Access - Other
   Status: FAIL
   ❌ Clients: Expected 4.0 | Actual 12
   ❌ Cases: Expected 2 | Actual 8

✅ Row 27: Partial Access - Cannot Be Severed
   Status: PASS
   ✅ Clients: Expected 114.0 | Actual 114
   ✅ Cases: Expected 75 | Actual 75

✅ Row 28: No Info Released - Part X Deny
   Status: PASS
   ✅ Clients: Expected 0.0 | Actual 0
   ✅ Cases: Expected 0 | Actual 0

✅ Row 29: No Info Released - Records Not Found
   Status: PASS
   ✅ Clients: Expected 5.0 | Actual 5
   ✅ Cases: Expected 5 | Actual 5

✅ Row 30: No Info Released - Part X Does Not Apply
   Status: PASS
   ✅ Clients: Expected 9.0 | Actual 9
   ✅ Cases: Expected 6 | Actual 6

❌ Row 31: No Info Released - Other
   Status: FAIL
   ❌ Clients: Expected 4.0 | Actual 5
   ❌ Cases: Expected 3 | Actual 4

✅ Row 32: No Info Released - Cannot Be Severed
   Status: PASS
   ✅ Clients: Expected 0.0 | Actual 0
   ✅ Cases: Expected 0 | Actual 0

✅ Row 33: No Information Released - Intake Only
   Status: PASS
   ✅ Participants: Expected 89.0 | Actual 89
   ✅ Intake Cases: Expected 62 | Actual 62

✅ Row 34: Withdrawn or Abandoned
   Status: PASS
   ✅ Clients: Expected 52.0 | Actual 52
   ✅ Cases: Expected 20 | Actual 20
   ✅ Participants: Expected 85.0 | Actual 85
   ✅ Intake Cases: Expected 46 | Actual 46

✅ Row 35: Documentation Completed
   Status: PASS
   ✅ Clients: Expected 131.0 | Actual 131
   ✅ Cases: Expected 92 | Actual 92

❌ Row 36: Total Distinct Outcomes
   Status: FAIL
   ❌ Clients: Expected 383.0 | Actual 506
   ❌ Cases: Expected 261 | Actual 343
   ✅ Participants: Expected 174.0 | Actual 174
   ✅ Intake Cases: Expected 108 | Actual 108

❌ Row 37: Partial/No Info - Part X Deny
   Status: FAIL
   ❌ Clients: Expected 0.0 | Actual 114
   ❌ Cases: Expected 0 | Actual 75

================================================================================
SUMMARY: 11 PASSED, 5 FAILED, 0 ERRORS
SUCCESS RATE: 68.8%
================================================================================
'''